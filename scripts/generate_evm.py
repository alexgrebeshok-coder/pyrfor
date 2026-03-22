#!/usr/bin/env python3
"""
CEOClaw EVM Model Generator
Generates Excel files with EVM (Earned Value Management) formulas
"""

import sys
import json
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment


def generate_evm_model(project_data):
    """Generate EVM Excel model from project data"""

    wb = Workbook()

    # ========== Sheet 1: Inputs ==========
    ws_inputs = wb.active
    ws_inputs.title = "Inputs"

    # Project info
    inputs = [
        ("Project Name", project_data.get('name', 'Project')),
        ("Budget (BAC), ₽", project_data.get('budget', 10000000)),
        ("Duration (months)", project_data.get('duration', 12)),
        ("Start Date", project_data.get('start_date', datetime.now().strftime('%Y-%m-%d'))),
        ("Project Manager", project_data.get('manager', '')),
        ("Status", "Active"),
    ]

    for row, (label, value) in enumerate(inputs, 1):
        ws_inputs.cell(row=row, column=1, value=label)
        ws_inputs.cell(row=row, column=1).font = Font(bold=True)
        ws_inputs.cell(row=row, column=2, value=value)
        if "Budget" in label:
            ws_inputs.cell(row=row, column=2).number_format = '#,##0 ₽'

    # ========== Sheet 2: EVM Calculations ==========
    ws_evm = wb.create_sheet("EVM")

    # Headers
    headers = [
        'Month', 'PV (Planned)', 'EV (Earned)', 'AC (Actual)',
        'SPI', 'CPI', 'CV', 'SV', 'EAC', 'ETC', 'VAC'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_evm.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    # Monthly data
    duration = project_data.get('duration', 12)
    budget = project_data.get('budget', 10000000)

    for month in range(1, duration + 1):
        row = month + 1
        ws_evm.cell(row=row, column=1, value=month)

        # PV (Plan Value) - linear distribution
        pv = budget / duration
        ws_evm.cell(row=row, column=2, value=pv)
        ws_evm.cell(row=row, column=2).number_format = '#,##0 ₽'

        # EV (Earned Value) - 90% of PV (example)
        ws_evm.cell(row=row, column=3, value=f'=B{row}*0.9')
        ws_evm.cell(row=row, column=3).number_format = '#,##0 ₽'

        # AC (Actual Cost) - 105% of EV (example)
        ws_evm.cell(row=row, column=4, value=f'=C{row}*1.05')
        ws_evm.cell(row=row, column=4).number_format = '#,##0 ₽'

        # SPI = EV / PV
        ws_evm.cell(row=row, column=5, value=f'=IF(B{row}>0, C{row}/B{row}, 0)')
        ws_evm.cell(row=row, column=5).number_format = '0.00'

        # CPI = EV / AC
        ws_evm.cell(row=row, column=6, value=f'=IF(D{row}>0, C{row}/D{row}, 0)')
        ws_evm.cell(row=row, column=6).number_format = '0.00'

        # CV = EV - AC (Cost Variance)
        ws_evm.cell(row=row, column=7, value=f'=C{row}-D{row}')
        ws_evm.cell(row=row, column=7).number_format = '#,##0 ₽'

        # SV = EV - PV (Schedule Variance)
        ws_evm.cell(row=row, column=8, value=f'=C{row}-B{row}')
        ws_evm.cell(row=row, column=8).number_format = '#,##0 ₽'

        # EAC = BAC / CPI
        ws_evm.cell(row=row, column=9, value=f'=IF(F{row}>0, Inputs!$B$2/F{row}, 0)')
        ws_evm.cell(row=row, column=9).number_format = '#,##0 ₽'

        # ETC = EAC - AC (cumulative)
        ws_evm.cell(row=row, column=10, value=f'=I{row}-D{row}')
        ws_evm.cell(row=row, column=10).number_format = '#,##0 ₽'

        # VAC = BAC - EAC (Variance at Completion)
        ws_evm.cell(row=row, column=11, value=f'=Inputs!$B$2-I{row}')
        ws_evm.cell(row=row, column=11).number_format = '#,##0 ₽'

    # Conditional formatting for SPI/CPI
    # Green if > 1.0, Red if < 1.0
    for row in range(2, duration + 2):
        # SPI
        cell = ws_evm.cell(row=row, column=5)
        if month > 1:
            pass  # Would add conditional formatting here

    # ========== Sheet 3: Scenarios ==========
    ws_scenarios = wb.create_sheet("Scenarios")

    ws_scenarios['A1'] = "Scenario Analysis"
    ws_scenarios['A1'].font = Font(bold=True, size=14)

    ws_scenarios['A3'] = "Scenario"
    ws_scenarios['B3'] = "Budget Factor"
    ws_scenarios['C3'] = "Duration Factor"
    ws_scenarios['D3'] = "Risk Factor"
    ws_scenarios['E3'] = "Adjusted Budget"

    for col in range(1, 6):
        ws_scenarios.cell(row=3, column=col).font = Font(bold=True)

    scenarios = [
        ('Best Case', 0.9, 0.9, 0.8),
        ('Base Case', 1.0, 1.0, 1.0),
        ('Worst Case', 1.2, 1.3, 1.5),
    ]

    for i, (name, budget_f, duration_f, risk_f) in enumerate(scenarios, 4):
        ws_scenarios.cell(row=i, column=1, value=name)
        ws_scenarios.cell(row=i, column=2, value=budget_f)
        ws_scenarios.cell(row=i, column=3, value=duration_f)
        ws_scenarios.cell(row=i, column=4, value=risk_f)
        ws_scenarios.cell(row=i, column=5, value=f'=Inputs!$B$2*B{i}*D{i}')
        ws_scenarios.cell(row=i, column=5).number_format = '#,##0 ₽'

    # ========== Sheet 4: Charts ==========
    ws_charts = wb.create_sheet("Charts")

    # S-Curve Chart
    chart = LineChart()
    chart.title = "S-Curve: PV vs EV vs AC"
    chart.style = 10
    chart.x_axis.title = "Month"
    chart.y_axis.title = "Value (₽)"
    chart.width = 15
    chart.height = 10

    # Add data from EVM sheet
    data = Reference(ws_evm, min_col=2, max_col=4, min_row=1, max_row=duration+1)
    chart.add_data(data, titles_from_data=True)

    ws_charts.add_chart(chart, "A1")

    # SPI/CPI Trend Chart
    chart2 = LineChart()
    chart2.title = "SPI & CPI Trend"
    chart2.style = 10
    chart2.x_axis.title = "Month"
    chart2.y_axis.title = "Index"
    chart2.width = 15
    chart2.height = 10

    data2 = Reference(ws_evm, min_col=5, max_col=6, min_row=1, max_row=duration+1)
    chart2.add_data(data2, titles_from_data=True)

    ws_charts.add_chart(chart2, "A20")

    # ========== Sheet 5: NPV/IRR ==========
    ws_npv = wb.create_sheet("NPV_IRR")

    ws_npv['A1'] = "Financial Analysis"
    ws_npv['A1'].font = Font(bold=True, size=14)

    ws_npv['A3'] = "Discount Rate"
    ws_npv['B3'] = 0.1  # 10%
    ws_npv['B3'].number_format = '0%'
    ws_npv['A3'].font = Font(bold=True)

    ws_npv['A5'] = "Year"
    ws_npv['B5'] = "Cash Flow"
    ws_npv['C5'] = "Discounted CF"

    for col in range(1, 4):
        ws_npv.cell(row=5, column=col).font = Font(bold=True)

    # NPV calculation (5 years)
    for year in range(1, 6):
        row = year + 5
        ws_npv.cell(row=row, column=1, value=year)

        # Simplified cash flow
        if year <= 2:
            cf = -budget / 2  # Investment phase
        else:
            cf = budget * 0.3  # Return phase

        ws_npv.cell(row=row, column=2, value=cf)
        ws_npv.cell(row=row, column=2).number_format = '#,##0 ₽'

        # Discounted cash flow
        ws_npv.cell(row=row, column=3, value=f'=B{row}/(1+$B$3)^A{row}')
        ws_npv.cell(row=row, column=3).number_format = '#,##0 ₽'

    ws_npv['A12'] = "NPV"
    ws_npv['B12'] = '=SUM(C6:C10)'
    ws_npv['B12'].number_format = '#,##0 ₽'
    ws_npv['B12'].font = Font(bold=True, size=12)
    ws_npv['A12'].font = Font(bold=True)

    ws_npv['A13'] = "IRR"
    ws_npv['B13'] = '=IRR(B6:B10)'
    ws_npv['B13'].number_format = '0.0%'
    ws_npv['B13'].font = Font(bold=True, size=12)
    ws_npv['A13'].font = Font(bold=True)

    # ========== Sheet 6: Summary Dashboard ==========
    ws_summary = wb.create_sheet("Dashboard")

    ws_summary['A1'] = f"EVM Dashboard: {project_data.get('name', 'Project')}"
    ws_summary['A1'].font = Font(bold=True, size=16)

    summary_data = [
        ("Budget (BAC)", f'=Inputs!B2'),
        ("Duration", f'=Inputs!B3 & " months"'),
        ("", ""),
        ("Current SPI", f'=EVM!E{duration+1}'),
        ("Current CPI", f'=EVM!F{duration+1}'),
        ("", ""),
        ("Estimate at Completion (EAC)", f'=EVM!I{duration+1}'),
        ("Estimate to Complete (ETC)", f'=EVM!J{duration+1}'),
        ("Variance at Completion (VAC)", f'=EVM!K{duration+1}'),
        ("", ""),
        ("NPV", '=NPV_IRR!B12'),
        ("IRR", '=NPV_IRR!B13'),
    ]

    for i, (label, formula) in enumerate(summary_data, 3):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=1).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=formula)
        if "Budget" in label or "EAC" in label or "ETC" in label or "VAC" in label or "NPV" in label:
            ws_summary.cell(row=i, column=2).number_format = '#,##0 ₽'
        elif "SPI" in label or "CPI" in label or "IRR" in label:
            ws_summary.cell(row=i, column=2).number_format = '0.00'

    # Save
    project_name = project_data.get('name', 'Project').replace(' ', '_').replace('/', '_')
    output_path = f"EVM_{project_name}.xlsx"
    wb.save(output_path)

    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        # Default demo project
        project_data = {
            'name': 'Demo_Project',
            'budget': 10000000,
            'duration': 12,
            'start_date': '2026-01-01',
            'manager': 'Иванов И.И.'
        }
    else:
        # Parse JSON from command line
        project_data = json.loads(sys.argv[1])

    output_file = generate_evm_model(project_data)
    print(output_file)


def generate_portfolio_evm(projects_data, output_path=None):
    """Generate EVM Excel for multiple projects"""
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Portfolio_Summary"
    
    # Headers
    headers = ['Project', 'Budget (BAC)', 'Progress %', 'Actual Cost', 'CPI', 'SPI', 'EAC', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    total_planned = 0
    total_actual = 0
    
    for idx, project in enumerate(projects_data.get('projects', []), 2):
        planned = project.get('budget_planned', 0)
        actual = project.get('budget_actual', 0)
        progress = project.get('progress', 0)
        
        # Calculate EVM metrics
        pv = planned * progress / 100  # Planned Value (Earned)
        ev = pv  # Assume EV = PV for simplicity
        cpi = (ev / actual) if actual > 0 else 1.0
        spi = 1.0  # Assume on schedule
        eac = planned / cpi if cpi > 0 else planned
        
        total_planned += planned
        total_actual += actual
        
        status = "On Track" if cpi >= 0.95 else ("At Risk" if cpi >= 0.85 else "Critical")
        
        ws_summary.cell(row=idx, column=1, value=project.get('name', f'Project {idx-1}'))
        ws_summary.cell(row=idx, column=2, value=planned)
        ws_summary.cell(row=idx, column=2).number_format = '#,##0 ₽'
        ws_summary.cell(row=idx, column=3, value=f'{progress}%')
        ws_summary.cell(row=idx, column=4, value=actual)
        ws_summary.cell(row=idx, column=4).number_format = '#,##0 ₽'
        ws_summary.cell(row=idx, column=5, value=round(cpi, 2))
        ws_summary.cell(row=idx, column=6, value=round(spi, 2))
        ws_summary.cell(row=idx, column=7, value=round(eac, 0))
        ws_summary.cell(row=idx, column=7).number_format = '#,##0 ₽'
        ws_summary.cell(row=idx, column=8, value=status)
    
    # Portfolio totals
    last_row = len(projects_data.get('projects', [])) + 2
    ws_summary.cell(row=last_row, column=1, value="TOTAL")
    ws_summary.cell(row=last_row, column=1).font = Font(bold=True)
    ws_summary.cell(row=last_row, column=2, value=total_planned)
    ws_summary.cell(row=last_row, column=2).number_format = '#,##0 ₽'
    ws_summary.cell(row=last_row, column=2).font = Font(bold=True)
    ws_summary.cell(row=last_row, column=4, value=total_actual)
    ws_summary.cell(row=last_row, column=4).number_format = '#,##0 ₽'
    ws_summary.cell(row=last_row, column=4).font = Font(bold=True)
    
    # Adjust column widths
    for col in range(1, 9):
        ws_summary.column_dimensions[chr(64+col)].width = 18
    
    # Save
    if output_path:
        wb.save(output_path)
        return output_path
    else:
        output_file = f"EVM_Portfolio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(output_file)
        return output_file


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate EVM Excel reports')
    parser.add_argument('--input', type=str, help='Path to input JSON file')
    parser.add_argument('--output', type=str, help='Path to output Excel file')
    parser.add_argument('--mode', type=str, default='portfolio', choices=['single', 'portfolio'],
                       help='Generation mode: single project or portfolio')
    
    args = parser.parse_args()
    
    if args.input:
        with open(args.input, 'r') as f:
            data = json.load(f)
        
        if args.mode == 'portfolio' or 'projects' in data:
            output = generate_portfolio_evm(data, args.output)
        else:
            # Single project mode
            project_data = data.get('projects', [data])[0] if 'projects' in data else data
            output = generate_evm_model(project_data)
            if args.output:
                import shutil
                shutil.move(output, args.output)
                output = args.output
    else:
        # Default demo
        project_data = {
            'name': 'Demo_Project',
            'budget': 10000000,
            'duration': 12,
            'start_date': '2026-01-01',
            'manager': 'Иванов И.И.'
        }
        output = generate_evm_model(project_data)
    
    print(output)
